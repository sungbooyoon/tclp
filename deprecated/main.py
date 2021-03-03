import numpy as np
import functions
import pareto_ga_multicrane as ga
import openpyxl
import time
import json
from tqdm import tqdm
import pickle

"""
INPUTS
"""
# T/C 설치 위치 및 유닛 설치 위치 불러오기
(
    crane_location_list,
    unit_location_list,
    trailer_location_list,
) = functions.get_location()
if any([value for value in crane_location_list if value in unit_location_list]):
    print("Locations of cranes and units overlap!")
crane_location_number = len(crane_location_list)
unit_location_number = len(unit_location_list)
trailer_location_number = len(trailer_location_list)

print(crane_location_list)

# 크레인 제원 불러오기
file_path = "data/simulation.json"
crane_data = functions.get_cranedb(file_path)

# Genetic algorithm parameters
POPULATION_NUMBER = 100
# 크레인 대수 3대로 설정되어 있음
CRANE_NUMBER = 4
MIN_NUMBER_PARENTS = 10
GENERATION = 200
MUTATION = 0.1

# Defining the population size.
pop_size = (POPULATION_NUMBER, crane_location_number)

start = time.time()
"""
OPTIMIZATION MODEL
"""
# Creating the initial population.
new_population = np.empty((0, crane_location_number), int)
while new_population.shape[0] < POPULATION_NUMBER:
    chromosome = functions.make_random_chromosome(crane_location_number, CRANE_NUMBER)
    if functions.check_chromosome(chromosome):
        new_population = np.append(new_population, [chromosome], axis=0)
        print(f"Created. {len(new_population)}")

crane_location_list = []
for i in new_population:
    crane_list = []
    crane_model = functions.get_crane(i)
    for x in range(len(crane_model)):
        crane_list.append([crane_model[x]["x"], crane_model[x]["y"]])
    crane_location_list.append(crane_list)
crane_location_list = np.array(crane_location_list)

best_outputs = []

for generation in tqdm(range(GENERATION)):
    # Measuring the fitness of each chromosome in the population.
    fitness = ga.cal_pop_fitness(new_population)
    pareto = functions.identify_pareto(fitness, new_population)
    best_outputs_pareto = ga.cal_pop_fitness(pareto)
    best_outputs_pareto_list = best_outputs_pareto.tolist()
    for i in best_outputs_pareto_list:
        best_outputs.append(i)
    # Selecting the best parents in the population for mating.
    parents = ga.select_mating_pool(new_population, pareto, MIN_NUMBER_PARENTS)
    offspring = np.empty((0, crane_location_number), int)
    offspring_number = pop_size[0] - parents.shape[0]
    while offspring.shape[0] < offspring_number:
        offspring_crossover = ga.crossover(
            parents, offspring_size=(offspring_number, crane_location_number)
        )
        # Adding some variations to the offspring using mutation.
        offspring_mutation = ga.mutation(offspring_crossover, crane_data, MUTATION)

        for chromosome in offspring_mutation:
            if functions.check_chromosome(chromosome):
                offspring = np.append(offspring, [chromosome], axis=0)
                if offspring.shape[0] == offspring_number:
                    break
            else:
                continue

    # Creating the new population based on the parents and offspring.
    new_population[0 : parents.shape[0], :] = parents
    new_population[parents.shape[0] :, :] = offspring

end = time.time()
print("Time: ", end - start)

# Getting the best solution after iterating finishing all generations.
fitness = ga.cal_pop_fitness(new_population)
pareto = functions.identify_pareto(fitness, new_population)
task_allocation = []
best_solution = []
location_list = []
time_list = []
space_list = []
for i in pareto:
    crane_model = functions.get_crane(i)
    location_list_i = functions.get_location_from_get_crane(i)
    best_solution.append(crane_model)
    location_list.append(location_list_i)
    task_allocation.append(functions.final_task_allocation(i))
    time_list.append(functions.get_total_operation_time(i))
    space_list.append(functions.get_collision_factor(crane_model))
# best_solution = np.array(best_solution)
best_outputs_pareto = ga.cal_pop_fitness(pareto)
best_outputs_pareto_list = best_outputs_pareto.tolist()
for i in best_outputs_pareto_list:
    best_outputs.append(i)
best_outputs = np.array(best_outputs)
x_all = fitness[:, 0]
y_all = fitness[:, 1]
x_pareto_all = best_outputs[:, 0]
y_pareto_all = best_outputs[:, 1]
range_pareto_all = np.array([i for i in range(len(best_outputs))])
x_pareto = best_outputs_pareto[:, 0]
y_pareto = best_outputs_pareto[:, 1]

all_data = [
    task_allocation,
    best_solution,
    location_list,
    time_list,
    space_list,
    best_outputs.tolist(),
]

with open("200_100.pickle", "wb") as f:
    pickle.dump(all_data, f)

# result = {}
# for i in range(len(x_pareto)):
#     result["idx"] = i
#     result["idx"]["tc_number"] = len(best_solution[i])
#     result["idx"]["model"] = best_solution[i]["model"]
#     result["idx"]["location"] = location_list[i]
#     result["idx"]["cost"] = x_pareto[i]
#     result["idx"]["duration"] = y_pareto[i]
#     result["idx"]["space"] = str(space_list[i])

# with open("../data/test.json", "w") as json_file:
#     json.dump(result, json_file)


"""
Outputs
"""
# for a in range(len(x_pareto_new)):
#     # 운용비용(원)
#     sheet.cell(row=a + 1, column=1).value = x_pareto[a]
#     # 운용기간(시간)
#     sheet.cell(row=a + 1, column=2).value = y_pareto[a]
#     # 최적 대수, 위치, 모델
#     sheet.cell(row=a + 1, column=3).value = str(best_solution[a])
#     # 간섭 면적(m2)
#     sheet.cell(row=a + 1, column=4).value = str(space_list[a])

# wb.save("test1.xlsx")
